from openpyxl import load_workbook
from openpyxl import Workbook
y1=[]
y2=[]
y3=[]
wb = load_workbook(filename='aaa.xlsx')
#print(wb.get_sheet_by_name)
ws = wb.active
a=ws.max_column
#print(a)
#c1=ws.cell('A1')
for i in range(2,22,1):
    y1.append(ws.cell(row=i,column=2).value) #更加方便实用
for i in range(2, 22, 1):
    y2.append(ws.cell(row=i, column=3).value)  # 更加方便实用
for i in range(2, 22, 1):
    y3.append(ws.cell(row=i, column=4).value)  # 更加方便实用
print(y1)