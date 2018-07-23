import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook
y1=[]
y2=[]
y3=[]
wb = load_workbook(filename='aaa.xlsx')
ws = wb.active
a=ws.max_column
for i in range(2,22,1):
    y1.append(ws.cell(row=i,column=2).value) #更加方便实用
for i in range(2, 22, 1):
    y2.append(ws.cell(row=i, column=3).value)  # 更加方便实用
for i in range(2, 22, 1):
    y3.append(ws.cell(row=i, column=4).value)  # 更加方便实用
#y1=[-4.80999,-5.81,-6.37014,-6.602042,-7.789,-8.0546,-8.842940,-9.736321,-12.347975]
x1=range(2,42,2)
#y1=[-4.80999,-5.81,-6.37014,-6.602042,-7.789,-8.0546,-8.842940,-9.736321,-12.347975]
#x1=[1.8,2.6,3.5,3.6,5.2,5.4,7,7.8,10.5]
#x2=range(0,10)
#y2=[5,8,0,30,20,40,50,10,40,15]
plt.plot(x1,y1,label='Direct Method',linewidth=1,color='b',marker='.',
markerfacecolor='blue',markersize=7)
plt.plot(x1,y2,label='Our Proposed Method',linewidth=1,color='r',marker='.',
markerfacecolor='red',markersize=7)
plt.plot(x1,y3,label='Traditional Large Signal Calibration Method',linewidth=1,color='g',marker='.',
markerfacecolor='green',markersize=7)
#plt.plot(x2,y2,label='second line')
plt.xlabel('Frequency(GHz)')
plt.ylabel('Magnitude(dB)')
plt.title('PortPower Mesurement')
plt.legend()
plt.show()

